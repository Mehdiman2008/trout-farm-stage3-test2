"""
fx.py — بنچمارک ارزی / ارزش زمانی سرمایه (سه‌ماهه)
====================================================
هدف: پیش‌بینی دلار نیست. فقط این سؤال:
    «اگر همین سرمایه ابتدای سه‌ماهه به دلار تبدیل می‌شد، الان چقدر بود؟»

روش دقیقاً مطابق specification:
    FX_start  = اولین نرخ موجود در ابتدای سه‌ماهه
    FX_end    = آخرین نرخ موجود در انتهای سه‌ماهه
    FX_return = FX_end / FX_start - 1
    USD_alternative_end_value = Capital_q × FX_end / FX_start

منبع داده: شیت «روزانه» و ستون «قیمت پایانی (تومان)» از فایل
TGJU_USD_3Y_Daily_Weekly_Close(1).xlsx در پوشه data/.

هیچ داده ساختگی/PLACEHOLDER ساخته نمی‌شود (اصلاح ۸). اگر فایل موجود
نباشد یا یک دوره داده نداشته باشد، همان دوره «Missing / Unavailable»
گزارش می‌شود.
"""
from __future__ import annotations


import os
from datetime import date, timedelta

MISSING_NOTE = ("فایل واقعی نرخ دلار در پوشه data/ یافت نشد. "
                "هیچ داده جایگزینی ساخته نمی‌شود؛ بنچمارک ارزی «در دسترس نیست» است.")


# ------------------------------------------------------------------ loading
def load_fx_into_db(db, A) -> dict:
    """تلاش برای خواندن فایل اکسل؛ در صورت نبود، ساخت سری جایگزین."""
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = A.get("fx.file")
    if not os.path.isabs(path):
        path = os.path.join(root, path)
    if os.path.exists(path):
        try:
            rows = _read_excel(path, A.get("fx.sheet"), A.get("fx.column"))
            if rows:
                db.fx_replace_all(rows, source="TGJU-excel")
                db.meta_set("fx_source", "file")
                db.meta_set("fx_file", os.path.basename(path))
                return {"ok": True, "source": "file", "rows": len(rows),
                        "file": os.path.basename(path)}
        except Exception as e:                                # pragma: no cover
            return _missing(db, note=f"خطا در خواندن فایل نرخ دلار: {e}")
    return _missing(db, note=MISSING_NOTE)


def _read_excel(path, sheet, column):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    header, rows = None, []
    for r in ws.iter_rows(values_only=True):
        if header is None:
            if r and any(c is not None for c in r):
                header = [str(c).strip() if c is not None else "" for c in r]
            continue
        rec = dict(zip(header, r))
        g = _find(rec, ["تاریخ میلادی", "date", "میلادی", "gregorian"])
        j = _find(rec, ["تاریخ شمسی", "شمسی", "jalali"])
        v = rec.get(column)
        if v is None:
            v = _find(rec, ["قیمت پایانی (تومان)", "تومان", "close"])
        if g is None or v is None:
            continue
        try:
            gd = g.date().isoformat() if hasattr(g, "date") else str(g)[:10]
            rows.append({"date_g": gd, "date_j": str(j) if j else None,
                         "close_toman": float(str(v).replace(",", ""))})
        except Exception:
            continue
    wb.close()
    rows.sort(key=lambda r: r["date_g"])
    return rows


def _find(rec, keys):
    for k, v in rec.items():
        kl = str(k).lower()
        for want in keys:
            if want.lower() in kl:
                return v
    return None


def _missing(db, note: str) -> dict:
    """هیچ سری ساختگی ساخته نمی‌شود؛ فقط وضعیت «داده موجود نیست» ثبت می‌گردد."""
    db.fx_replace_all([], source="missing")
    db.meta_set("fx_source", "missing")
    db.meta_set("fx_note", note)
    return {"ok": False, "source": "missing", "rows": 0, "note": note}


# ---------------------------------------------------------------- quarters
def _quarter_of(dt: date):
    return (dt.year, (dt.month - 1) // 3 + 1)


def _quarter_bounds(y, q):
    m0 = 3 * (q - 1) + 1
    start = date(y, m0, 1)
    end = date(y + (1 if q == 4 else 0), 1 if q == 4 else m0 + 3, 1) - timedelta(days=1)
    return start, end


class FXBenchmark:
    def __init__(self, db, A):
        self.db, self.A = db, A
        self.series = db.fx_series()
        self.source = db.meta_get("fx_source", "unknown")
        self.note = db.meta_get("fx_note")
        self.by_date = {r["date_g"]: r["close_toman"] for r in self.series}

    def latest(self):
        return self.series[-1] if self.series else None

    def _first_in(self, a: date, b: date):
        for r in self.series:
            if a.isoformat() <= r["date_g"] <= b.isoformat():
                return r
        return None

    def _last_in(self, a: date, b: date):
        found = None
        for r in self.series:
            if a.isoformat() <= r["date_g"] <= b.isoformat():
                found = r
        return found

    def quarters(self, capital_by_quarter: dict, as_of: date, n: int = 6) -> list:
        """
        capital_by_quarter: {(year,quarter): capital_toman}
        خروجی: لیست سه‌ماهه‌ها با FX return و ارزش جایگزین دلاری.
        """
        share = float(self.A.get("fx.benchmark_share"))
        y, q = _quarter_of(as_of)
        qs = []
        for _ in range(n):
            qs.append((y, q))
            q -= 1
            if q == 0:
                q = 4
                y -= 1
        qs.reverse()
        out = []
        for (yy, qq) in qs:
            a, b = _quarter_bounds(yy, qq)
            b_eff = min(b, as_of)
            if b_eff < a:
                continue
            r0 = self._first_in(a, b_eff)
            r1 = self._last_in(a, b_eff)
            if not r0 or not r1:
                # داده واقعی برای این دوره موجود نیست — ساخته نمی‌شود
                out.append({"year": yy, "quarter": qq, "label": f"{yy}-Q{qq}",
                            "start": a.isoformat(), "end": b_eff.isoformat(),
                            "available": False,
                            "capital": float(capital_by_quarter.get((yy, qq), 0.0)),
                            "note": "داده نرخ دلار برای این دوره موجود نیست"})
                continue
            fx0, fx1 = r0["close_toman"], r1["close_toman"]
            ret = fx1 / fx0 - 1 if fx0 else 0.0
            cap = float(capital_by_quarter.get((yy, qq), 0.0))
            benched = cap * share
            out.append({
                "available": True,
                "year": yy, "quarter": qq, "label": f"{yy}-Q{qq}",
                "start": a.isoformat(), "end": b_eff.isoformat(),
                "fx_start": fx0, "fx_end": fx1, "fx_return": ret,
                "capital": cap, "benchmark_share": share,
                "usd_alternative_end_value": benched * (fx1 / fx0) if fx0 else 0.0,
                "usd_alternative_gain": benched * ret,
                "is_current": (yy, qq) == _quarter_of(as_of),
                "partial": b_eff < b,
            })
        return out

    def current_quarter_kpi(self, capital: float, as_of: date,
                            farm_value_change: float | None = None) -> dict:
        y, q = _quarter_of(as_of)
        a, b = _quarter_bounds(y, q)
        b_eff = min(b, as_of)
        r0 = self._first_in(a, b_eff)
        r1 = self._last_in(a, b_eff)
        share = float(self.A.get("fx.benchmark_share"))
        if not r0 or not r1:
            return {"available": False, "source": self.source,
                    "note": self.note or "داده نرخ دلار برای این سه‌ماهه موجود نیست",
                    "missing": True, "label": f"{y}-Q{q}"}
        fx0, fx1 = r0["close_toman"], r1["close_toman"]
        ret = fx1 / fx0 - 1 if fx0 else 0.0
        benched = capital * share
        gain = benched * ret
        out = {
            "available": True,
            "source": self.source,
            "note": self.note if self.source != "file" else None,
            "missing": False,
            "file": self.db.meta_get("fx_file"),
            "label": f"{y}-Q{q}",
            "quarter_start": a.isoformat(), "as_of": b_eff.isoformat(),
            "fx_start": fx0, "fx_latest": fx1, "fx_return": ret,
            "capital": capital, "benchmark_share": share,
            "usd_alternative_end_value": benched * (fx1 / fx0) if fx0 else 0.0,
            "usd_alternative_gain": gain,
        }
        if farm_value_change is not None:
            out["farm_value_change"] = farm_value_change
            out["excess_over_fx"] = farm_value_change - gain
        return out
